from decimal import Decimal, InvalidOperation
from django.shortcuts import render, redirect, get_object_or_404
from .models import Purchase, PurchaseItem, Mill, Product, Payment, SaleItem, Sale, Broker
from django.db import transaction
from django.db.models import Q, Sum, F, FloatField, ExpressionWrapper
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime
import json
from django.utils import timezone
from io import BytesIO
from django.conf import settings
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib import colors
import qrcode
from reportlab.lib.utils import ImageReader
from num2words import num2words
from reportlab.pdfbase.pdfmetrics import stringWidth


def dashboard(request):
    return render(request, 'core/dashboard.html')

def add_purchase(request):
    mills = Mill.objects.all().order_by("mill_name")
    products = Product.objects.filter(is_active=True).order_by("rice_name")

    if request.method == "POST":
        mill_id = request.POST.get("mill")
        invoice_no = request.POST.get("invoice_no")
        purchase_date = request.POST.get("purchase_date")

        # multiple row values
        product_ids = request.POST.getlist("product[]")
        bag_weights = request.POST.getlist("bag_weight[]")
        bag_counts = request.POST.getlist("bag_count[]")
        rates = request.POST.getlist("purchase_price[]")  # ✅ rate per KG

        total_amount = 0

        with transaction.atomic():
            purchase = Purchase.objects.create(
                mill_id=mill_id,
                invoice_no=invoice_no,
                purchase_date=purchase_date,
                total_amount=0
            )

            for i in range(len(product_ids)):
                if not product_ids[i]:
                    continue

                bw = int(bag_weights[i] or 0)        # ✅ bag weight (50 default)
                bc = int(bag_counts[i] or 0)         # ✅ bag count
                rate = float(rates[i] or 0)          # ✅ rate per KG

                # ✅ Rule B: amount = bags * bag_weight * rate_per_kg
                line_total = bc * bw * rate
                total_amount += line_total

                PurchaseItem.objects.create(
                    purchase=purchase,
                    product_id=product_ids[i],
                    bag_weight=bw,
                    bag_count=bc,
                    purchase_price=rate
                )

            purchase.total_amount = total_amount
            purchase.save()

        messages.success(request, "✅ Purchase saved successfully!")
        return redirect("purchase_list")

    return render(request, "core/add_purchase.html", {"mills": mills, "products": products})


def purchase_list(request):
    purchases = (
        Purchase.objects.select_related("mill")
        .annotate(
            total_bags=Sum("purchaseitem__bag_count"),
            total_kg=Sum(
                ExpressionWrapper(
                    F("purchaseitem__bag_count") * F("purchaseitem__bag_weight"),
                    output_field=FloatField()
                )
            )
        )
        .order_by("-id")
    )

    return render(request, "core/purchase_list.html", {"purchases": purchases})

def purchase_detail(request, purchase_id):
    purchase = get_object_or_404(Purchase, id=purchase_id)
    items = PurchaseItem.objects.filter(purchase=purchase).select_related("product")

    # ✅ build rows with calculations
    item_rows = []
    total_amount = 0
    for it in items:
        row_kg = (it.bag_weight or 0) * (it.bag_count or 0)
        amount = row_kg * float(it.purchase_price or 0)
        total_amount += amount

        item_rows.append({
            "rice_name": it.product.rice_name,
            "bag_weight": it.bag_weight,
            "bag_count": it.bag_count,
            "row_kg": row_kg,
            "rate_per_kg": it.purchase_price,
            "amount": amount,
        })

    invoice_payments = Payment.objects.filter(
        related_type="purchase",
        purchase=purchase
    ).order_by("-payment_date", "-id")

    invoice_paid = invoice_payments.aggregate(s=Sum("amount"))["s"] or 0
    invoice_due = float(purchase.total_amount) - float(invoice_paid)

    total_bags = items.aggregate(s=Sum("bag_count"))["s"] or 0
    total_kg = items.aggregate(s=Sum(F("bag_weight") * F("bag_count")))["s"] or 0

    return render(request, "core/purchase_detail.html", {
        "purchase": purchase,
        "item_rows": item_rows,   # ✅ new
        "total_bags": total_bags,
        "total_kg": total_kg,
        "total_amount": total_amount,  # ✅ new
        "invoice_payments": invoice_payments,
        "invoice_paid": invoice_paid,
        "invoice_due": invoice_due,
    })


def edit_purchase(request, purchase_id):
    purchase = get_object_or_404(Purchase, id=purchase_id)
    mills = Mill.objects.all().order_by("mill_name")
    products = Product.objects.filter(is_active=True).order_by("rice_name")

    items = PurchaseItem.objects.filter(purchase=purchase).select_related("product")

    if request.method == "POST":
        mill_id = request.POST.get("mill")
        invoice_no = request.POST.get("invoice_no")
        purchase_date = request.POST.get("purchase_date")

        product_ids = request.POST.getlist("product[]")
        bag_weights = request.POST.getlist("bag_weight[]")
        bag_counts = request.POST.getlist("bag_count[]")
        rates = request.POST.getlist("purchase_price[]")

        total_amount = 0

        with transaction.atomic():
            # Update purchase header
            purchase.mill_id = mill_id
            purchase.invoice_no = invoice_no
            purchase.purchase_date = purchase_date
            purchase.save()

            # Remove old items
            PurchaseItem.objects.filter(purchase=purchase).delete()

            # Insert new items
            for i in range(len(product_ids)):
                if not product_ids[i]:
                    continue

                bw = int(bag_weights[i] or 0)
                bc = int(bag_counts[i] or 0)
                rate = float(rates[i] or 0)

                line_total = bc * bw * rate
                total_amount += line_total

                PurchaseItem.objects.create(
                    purchase=purchase,
                    product_id=product_ids[i],
                    bag_weight=bw,
                    bag_count=bc,
                    purchase_price=rate
                )

            purchase.total_amount = total_amount
            purchase.save()

        messages.success(request, "✅ Purchase updated successfully!")
        return redirect("purchase_detail", purchase_id=purchase.id)

    return render(request, "core/edit_purchase.html", {
        "purchase": purchase,
        "mills": mills,
        "products": products,
        "items": items,
    })

def delete_purchase(request, purchase_id):
    purchase = get_object_or_404(Purchase.objects.select_related("mill"), id=purchase_id)

    if request.method == "POST":
        purchase.delete()  # ✅ automatically deletes PurchaseItem also
        messages.success(request, "🗑 Purchase invoice deleted successfully!")
        return redirect("purchase_list")

    return render(request, "core/delete_purchase.html", {"purchase": purchase})


def add_mill(request):
    if request.method == "POST":
        Mill.objects.create(
            mill_name=request.POST.get("mill_name"),
            owner_name=request.POST.get("owner_name", ""),
            mobile=request.POST.get("mobile"),
            address=request.POST.get("address", ""),
            gst_number=request.POST.get("gst_number", ""),
            opening_balance=request.POST.get("opening_balance") or 0,
        )
        messages.success(request, "✅ Mill saved successfully!")
        return redirect("mill_list")

    return render(request, "core/add_mill.html")

def mill_list(request):
    q = request.GET.get("q", "").strip()

    mills = Mill.objects.all().order_by("-created_at")

    if q:
        mills = mills.filter(
            Q(mill_name__icontains=q) |
            Q(owner_name__icontains=q) |
            Q(mobile__icontains=q)
        )

    return render(request, "core/mill_list.html", {"mills": mills})


def edit_mill(request, mill_id):
    mill = get_object_or_404(Mill, id=mill_id)

    if request.method == "POST":
        mill.mill_name = request.POST.get("mill_name")
        mill.owner_name = request.POST.get("owner_name", "")
        mill.mobile = request.POST.get("mobile")
        mill.address = request.POST.get("address", "")
        mill.gst_number = request.POST.get("gst_number", "")
        mill.opening_balance = request.POST.get("opening_balance") or 0
        mill.save()

        messages.success(request, "Mill updated successfully ✅")
        return redirect("mill_list")

    return render(request, "core/edit_mill.html", {"mill": mill})


def delete_mill(request, mill_id):
    mill = get_object_or_404(Mill, id=mill_id)

    if request.method == "POST":
        mill.delete()
        messages.success(request, "Mill deleted successfully 🗑️")
        return redirect("mill_list")

    return render(request, "core/delete_mill.html", {"mill": mill})


def product_list(request):
    q = request.GET.get("q", "").strip()

    products = Product.objects.all().order_by("-id")  # ✅ no created_at in your model

    if q:
        products = products.filter(
            Q(rice_name__icontains=q) |
            Q(hsn_code__icontains=q)
        )

    return render(request, "core/product_list.html", {"products": products})


def add_product(request):
    if request.method == "POST":
        Product.objects.create(
            rice_name=request.POST.get("rice_name"),
            hsn_code=request.POST.get("hsn_code", ""),
            gst_percent=request.POST.get("gst_percent") or 0,
            is_active=True if request.POST.get("is_active") == "on" else False,
        )
        messages.success(request, "✅ Product saved successfully!")
        return redirect("product_list")

    return render(request, "core/add_product.html")


def edit_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == "POST":
        product.rice_name = request.POST.get("rice_name")
        product.hsn_code = request.POST.get("hsn_code", "")
        product.gst_percent = request.POST.get("gst_percent") or 0
        product.is_active = True if request.POST.get("is_active") == "on" else False
        product.save()

        messages.success(request, "✅ Product updated successfully!")
        return redirect("product_list")

    return render(request, "core/edit_product.html", {"product": product})


def delete_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == "POST":
        product.delete()
        messages.success(request, "🗑 Product deleted successfully!")
        return redirect("product_list")

    return render(request, "core/delete_product.html", {"product": product})


def product_report(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    total_purchase_bags = (
        PurchaseItem.objects.filter(product=product)
        .aggregate(s=Sum("bag_count"))["s"] or 0
    )

    total_sale_bags = (
        SaleItem.objects.filter(product=product)
        .aggregate(s=Sum("bag_count"))["s"] or 0
    )

    current_stock_bags = total_purchase_bags - total_sale_bags

    purchase_items = (
        PurchaseItem.objects.select_related("purchase", "purchase__mill")
        .filter(product=product)
        .order_by("-purchase__purchase_date")
    )

    return render(request, "core/product_report.html", {
        "product": product,
        "total_purchase_bags": total_purchase_bags,
        "total_sale_bags": total_sale_bags,
        "current_stock_bags": current_stock_bags,
        "purchase_items": purchase_items,
    })



def mill_report_detail(request, mill_id):
    mill = get_object_or_404(Mill, id=mill_id)

    # ✅ all purchases of this mill
    purchases = Purchase.objects.filter(mill=mill).order_by("-purchase_date", "-id")

    # ✅ all payments to this mill (direct + invoice wise)
    payments = Payment.objects.filter(
        related_type="purchase",
        mill=mill
    ).order_by("-payment_date", "-id")

    # ✅ totals
    total_purchase = purchases.aggregate(s=Sum("total_amount"))["s"] or 0
    total_paid = payments.aggregate(s=Sum("amount"))["s"] or 0

    balance = float(mill.opening_balance) + float(total_purchase) - float(total_paid)

    # ✅ grand totals (bags + kg across all purchases)
    all_items = PurchaseItem.objects.filter(purchase__mill=mill)

    grand_total_bags = all_items.aggregate(s=Sum("bag_count"))["s"] or 0
    grand_total_kg = all_items.aggregate(s=Sum(F("bag_weight") * F("bag_count")))["s"] or 0

    # ✅ build purchase rows (each invoice)
    purchase_rows = []
    for p in purchases:
        items = PurchaseItem.objects.filter(purchase=p)

        total_bags = items.aggregate(s=Sum("bag_count"))["s"] or 0
        total_kg = items.aggregate(s=Sum(F("bag_weight") * F("bag_count")))["s"] or 0

        # invoice-wise paid
        paid = Payment.objects.filter(
            related_type="purchase",
            purchase=p
        ).aggregate(s=Sum("amount"))["s"] or 0

        due = float(p.total_amount) - float(paid)
        if due < 0:
            due = 0

        avg_rate = 0
        if total_kg:
            avg_rate = float(p.total_amount) / float(total_kg)

        purchase_rows.append({
            "id": p.id,
            "purchase_date": p.purchase_date,
            "invoice_no": p.invoice_no,

            "total_bags": total_bags,
            "total_kg": total_kg,
            "avg_rate": round(avg_rate, 2),

            "total_amount": p.total_amount,
            "paid": paid,
            "due": round(due, 2),
        })

    return render(request, "core/mill_report_detail.html", {
        "mill": mill,

        "purchases": purchases,
        "payments": payments,

        "total_purchase": total_purchase,
        "total_paid": total_paid,
        "balance": round(balance, 2),

        "grand_total_bags": grand_total_bags,
        "grand_total_kg": grand_total_kg,

        "purchase_rows": purchase_rows,
    })


def add_mill_payment(request, mill_id):
    mill = get_object_or_404(Mill, id=mill_id)

    if request.method == "POST":
        amount = request.POST.get("amount")
        payment_mode = request.POST.get("payment_mode")
        payment_date = request.POST.get("payment_date")
        notes = request.POST.get("notes", "")

        Payment.objects.create(
            related_type="purchase",
            mill=mill,          # ✅ direct mill payment
            purchase=None,      # ✅ not invoice linked
            amount=amount,
            payment_mode=payment_mode,
            payment_date=payment_date,
            notes=notes
        )

        messages.success(request, "✅ Payment saved successfully!")
        return redirect("mill_report_detail", mill_id=mill.id)

    return render(request, "core/add_mill_payment.html", {"mill": mill})


def add_purchase_payment(request, purchase_id):
    purchase = get_object_or_404(Purchase, id=purchase_id)
    mill = purchase.mill

    if request.method == "POST":
        amount = request.POST.get("amount")
        payment_mode = request.POST.get("payment_mode")
        payment_date = request.POST.get("payment_date")
        notes = request.POST.get("notes", "")

        Payment.objects.create(
            related_type="purchase",
            mill=mill,
            purchase=purchase,   # ✅ invoice linked
            amount=amount,
            payment_mode=payment_mode,
            payment_date=payment_date,
            notes=notes
        )

        messages.success(request, "✅ Purchase invoice payment saved!")
        return redirect("purchase_detail", purchase_id=purchase.id)

    return render(request, "core/add_purchase_payment.html", {
        "purchase": purchase,
        "mill": mill
    })

def mill_report_excel(request, mill_id):
    mill = get_object_or_404(Mill, id=mill_id)

    # reuse same data from your mill_report_detail logic
    purchases = Purchase.objects.filter(mill=mill).order_by("-purchase_date", "-id")
    payments = Payment.objects.filter(related_type="purchase", mill=mill).order_by("-payment_date", "-id")

    total_purchase = purchases.aggregate(s=Sum("total_amount"))["s"] or 0
    total_paid = payments.aggregate(s=Sum("amount"))["s"] or 0
    balance = float(mill.opening_balance) + float(total_purchase) - float(total_paid)

    all_items = PurchaseItem.objects.filter(purchase__mill=mill)
    grand_total_bags = all_items.aggregate(s=Sum("bag_count"))["s"] or 0
    grand_total_kg = all_items.aggregate(s=Sum(F("bag_weight") * F("bag_count")))["s"] or 0

    # Purchase rows (invoice wise)
    purchase_rows = []
    for p in purchases:
        items = PurchaseItem.objects.filter(purchase=p)
        total_bags = items.aggregate(s=Sum("bag_count"))["s"] or 0
        total_kg = items.aggregate(s=Sum(F("bag_weight") * F("bag_count")))["s"] or 0

        paid = Payment.objects.filter(related_type="purchase", purchase=p).aggregate(s=Sum("amount"))["s"] or 0
        due = float(p.total_amount) - float(paid)
        if due < 0:
            due = 0

        avg_rate = round(float(p.total_amount) / float(total_kg), 2) if total_kg else 0

        purchase_rows.append([str(p.purchase_date), p.invoice_no, total_bags, total_kg, avg_rate, float(p.total_amount), float(paid), float(due)])

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Mill Report"

    # Header
    ws.append(["Mill Report"])
    ws.append([f"Mill: {mill.mill_name}"])
    ws.append([f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"])
    ws.append([])

    # Summary
    ws.append(["Opening Balance", float(mill.opening_balance)])
    ws.append(["Total Purchase", float(total_purchase)])
    ws.append(["Total Paid", float(total_paid)])
    ws.append(["Balance Due", float(balance)])
    ws.append(["Total Bags Purchased", grand_total_bags])
    ws.append(["Total KG Purchased", float(grand_total_kg)])
    ws.append([])

    # Purchases table
    ws.append(["Date", "Invoice", "Bags", "KG", "Avg Rate/KG", "Amount", "Paid", "Due"])
    for row in purchase_rows:
        ws.append(row)

    # Adjust column width
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Response
    # filename = f"Mill_Report_{mill.mill_name.replace(' ', '_')}.xlsx"
    dt = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"{mill.mill_name.strip().replace(' ', '_')}_{dt}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def mill_report_pdf(request, mill_id):
    mill = get_object_or_404(Mill, id=mill_id)

    purchases = Purchase.objects.filter(mill=mill).order_by("purchase_date", "id")
    payments = Payment.objects.filter(
        related_type="purchase",
        mill=mill
    ).order_by("payment_date", "id")

    # totals
    total_purchase = purchases.aggregate(s=Sum("total_amount"))["s"] or 0
    total_paid = payments.aggregate(s=Sum("amount"))["s"] or 0
    balance = float(mill.opening_balance) + float(total_purchase) - float(total_paid)

    all_items = PurchaseItem.objects.filter(purchase__mill=mill)
    grand_total_bags = all_items.aggregate(s=Sum("bag_count"))["s"] or 0
    grand_total_kg = all_items.aggregate(s=Sum(F("bag_weight") * F("bag_count")))["s"] or 0

    # response
    dt = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"{mill.mill_name.strip().replace(' ', '_')}_{dt}.pdf"
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    elements = []

    # ---------- TITLE ----------
    elements.append(Paragraph("<b>Maa Bagwati Bhandar</b>", styles["Title"]))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f"<b>Mill:</b> {mill.mill_name}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    # ---------- SUMMARY TABLE ----------
    summary_data = [
        ["Opening Balance", f"Rs {mill.opening_balance}"],
        ["Total Purchase", f"Rs {total_purchase}"],
        ["Total Paid", f"Rs {total_paid}"],
        ["Balance Due", f"Rs {round(balance, 2)}"],
        ["Total Bags Purchased", str(grand_total_bags)],
        ["Total KG Purchased", str(grand_total_kg)],
    ]

    summary_table = Table(summary_data, colWidths=[220, 250])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
    ]))

    elements.append(Paragraph("<b>Summary</b>", styles["Heading2"]))
    elements.append(summary_table)
    elements.append(Spacer(1, 14))

    # ---------- PURCHASES TABLE ----------
    elements.append(Paragraph("<b>Purchases (Invoice-wise)</b>", styles["Heading2"]))

    purchase_data = [[
        "Date", "Invoice", "Bags", "KG", "Rate/KG", "Amount", "Paid", "Due"
    ]]

    for p in purchases:
        items = PurchaseItem.objects.filter(purchase=p)
        total_bags = items.aggregate(s=Sum("bag_count"))["s"] or 0
        total_kg = items.aggregate(s=Sum(F("bag_weight") * F("bag_count")))["s"] or 0

        paid = Payment.objects.filter(related_type="purchase", purchase=p).aggregate(s=Sum("amount"))["s"] or 0
        due = float(p.total_amount) - float(paid)
        if due < 0:
            due = 0

        rate = round(float(p.total_amount) / float(total_kg), 2) if total_kg else 0

        purchase_data.append([
            str(p.purchase_date),
            p.invoice_no,
            str(total_bags),
            str(total_kg),
            f"Rs {rate}",
            f"Rs {p.total_amount}",
            f"Rs {paid}",
            f"Rs {round(due, 2)}"
        ])

        # ✅ Multiple payments under same invoice
        inv_pays = Payment.objects.filter(related_type="purchase", purchase=p).order_by("payment_date", "id")

        if inv_pays.exists():
            purchase_data.append(["", "", "", "", "", "", "", ""])  # empty row
            purchase_data.append(["", "Payments for Invoice:", "", "", "", "", "", ""])

            for pay in inv_pays:
                purchase_data.append([
                    "",
                    f"- {pay.payment_date} ({pay.payment_mode})",
                    "",
                    "",
                    "",
                    "",
                    f"Rs {pay.amount}",
                    pay.notes or ""
                ])

            purchase_data.append(["", "", "", "", "", "", "", ""])  # separator row

    purchase_table = Table(purchase_data, repeatRows=1, colWidths=[65, 85, 45, 45, 60, 70, 60, 70])
    purchase_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    elements.append(purchase_table)
    elements.append(Spacer(1, 14))

    # ---------- PAYMENTS TABLE ----------
    elements.append(PageBreak())
    elements.append(Paragraph("<b>All Payments (Direct + Invoice-wise)</b>", styles["Heading2"]))

    pay_data = [["Date", "Mode", "Invoice", "Amount"]]

    for pay in payments:
        invoice_txt = "-"
        if pay.purchase:
            invoice_txt = pay.purchase.invoice_no

        pay_data.append([
            str(pay.payment_date),
            pay.payment_mode,
            invoice_txt,
            f"Rs {pay.amount}"
        ])

    pay_table = Table(pay_data, repeatRows=1, colWidths=[70, 70, 90, 70, 220])
    pay_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (3, 1), (3, -1), "RIGHT"),
    ]))

    elements.append(pay_table)

    # build PDF
    doc.build(elements)
    return response


def sale_list(request):
    q = request.GET.get("q", "")
    sales = Sale.objects.all().order_by("-sale_date", "-id")

    if q:
        sales = sales.filter(customer_name__icontains=q)

    return render(request, "core/sale_list.html", {"sales": sales, "q": q})

def sale_list(request):
    q = request.GET.get("q", "").strip()

    sales = Sale.objects.select_related("broker").all().order_by("-sale_date", "-id")

    if q:
        sales = sales.filter(
            Q(customer_name__icontains=q) |
            Q(invoice_no__icontains=q) |
            Q(broker__broker_name__icontains=q)
        )

    rows = []
    for s in sales:
        paid = Payment.objects.filter(
            related_type="sale",
            sale=s
        ).aggregate(x=Sum("amount"))["x"] or 0

        due = float(s.total_amount) - float(paid)

        if due <= 0:
            status = "PAID"
        elif paid > 0:
            status = "PARTIAL"
        else:
            status = "DUE"

        rows.append({
            "id": s.id,
            "sale_date": s.sale_date,
            "invoice_no": s.invoice_no,
            "customer_name": s.customer_name,
            "broker": s.broker,
            "total_amount": s.total_amount,
            "paid": paid,
            "due": round(due, 2),
            "status": status,
        })

    return render(request, "core/sale_list.html", {
        "rows": rows,
        "q": q,
    })


from django.db.models import Sum

def sale_detail(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id)

    # internal breakup items (BUY cost)
    items = SaleItem.objects.filter(sale=sale).select_related("mill", "product")

    # Rice payments only
    payments = Payment.objects.filter(
        related_type="sale",
        sale=sale
    ).order_by("-payment_date", "-id")

    # ✅ selling side totals (rice invoice)
    rice_total = float(sale.taxable_amount) + float(sale.gst_amount)

    # total kg from sale header
    total_kg = float(sale.total_quantity_kg or 0)

    # ✅ selling rate per kg (auto)
    selling_rate_per_kg = 0
    if total_kg > 0:
        selling_rate_per_kg = float(sale.taxable_amount) / total_kg

    # ✅ total bags from breakup rows
    total_bags = items.aggregate(s=Sum("bag_count"))["s"] or 0

    # ✅ rice received = advance + other payments
    paid_extra = payments.aggregate(s=Sum("amount"))["s"] or 0
    rice_received_total = float(sale.advance_received) + float(paid_extra)
    rice_due = rice_total - rice_received_total
    if rice_due < 0:
        rice_due = 0

    # ✅ transport due (separate)
    transport_due = float(sale.transport_charge) - (
        float(sale.transport_paid_by_dealer) + float(sale.transport_paid_by_customer)
    )
    if transport_due < 0:
        transport_due = 0

    # ✅ buy cost total from breakup rows
    buy_cost_total = items.aggregate(s=Sum("amount"))["s"] or 0

    # ✅ profit estimate (rice selling - buy cost)
    profit_estimate = rice_total - float(buy_cost_total)

    return render(request, "core/sale_detail.html", {
        "sale": sale,
        "items": items,

        # Rice invoice (selling)
        "total_bags": total_bags,
        "total_kg": round(total_kg, 2),
        "selling_rate_per_kg": round(selling_rate_per_kg, 2),
        "rice_total": round(rice_total, 2),
        "paid_extra": round(float(paid_extra), 2),
        "rice_received_total": round(rice_received_total, 2),
        "rice_due": round(rice_due, 2),

        # Transport
        "transport_due": round(transport_due, 2),

        # Internal (buying)
        "buy_cost_total": round(float(buy_cost_total), 2),
        "profit_estimate": round(float(profit_estimate), 2),

        # payments table
        "payments": payments,
    })

def add_sale_payment(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id)

    if request.method == "POST":
        amount = request.POST.get("amount")
        payment_mode = request.POST.get("payment_mode")
        payment_date = request.POST.get("payment_date")
        notes = request.POST.get("notes", "")

        Payment.objects.create(
            related_type="sale",
            sale=sale,
            amount=amount,
            payment_mode=payment_mode,
            payment_date=payment_date,
            notes=notes
        )

        messages.success(request, "✅ Sale payment added!")
        return redirect("sale_detail", sale_id=sale.id)

    return render(request, "core/add_sale_payment.html", {"sale": sale})

def generate_sale_invoice_no():
    # Example: SAL-20260118-0001
    today = datetime.now().strftime("%Y%m%d")
    last = Sale.objects.filter(invoice_no__startswith=f"SAL-{today}").order_by("-id").first()
    if last and last.invoice_no:
        try:
            last_seq = int(last.invoice_no.split("-")[-1])
        except:
            last_seq = 0
    else:
        last_seq = 0
    return f"SAL-{today}-{last_seq+1:04d}"


# def add_sale(request):
    products = Product.objects.filter(is_active=True).order_by("rice_name")
    brokers = Broker.objects.all().order_by("broker_name")

    # Purchase stock list for dropdown
    purchase_items = (
        PurchaseItem.objects
        .select_related("purchase", "purchase__mill", "product")
        .order_by("-purchase__purchase_date", "-id")
    )

    purchase_items_json = json.dumps([
        {
            "id": pi.id,
            "product_id": pi.product_id,
            "bag_weight": pi.bag_weight,
            "label": f"{pi.purchase.invoice_no} / {pi.purchase.mill.mill_name} / Buy ₹{pi.purchase_price}/KG / {pi.bag_weight}kg bag",
        }
        for pi in purchase_items
    ])

    # -------------------------
    # GET → show form
    # -------------------------
    if request.method == "GET":
        return render(request, "core/add_sale.html", {
            "products": products,
            "brokers": brokers,
            "purchase_items_json": purchase_items_json,
        })

    # -------------------------
    # POST → review or save
    # -------------------------
    step = request.POST.get("step", "review")  # "review" or "save"

    # Common fields
    sale_date = request.POST.get("sale_date")
    customer_name = request.POST.get("customer_name")
    customer_gst = request.POST.get("customer_gst", "")
    broker_id = request.POST.get("broker_id") or None

    vehicle_number = request.POST.get("vehicle_number")
    driver_name = request.POST.get("driver_name")
    transporter_name = request.POST.get("transporter_name")

    # Rice selling fields
    try:
        product_id = int(request.POST.get("product_id") or 0)
    except:
        product_id = 0

    bag_weight = Decimal(request.POST.get("bag_weight") or "0")
    total_bags = int(request.POST.get("total_bags") or 0)

    rate_per_kg = Decimal(request.POST.get("rate_per_kg") or "0")
    gst_percent = Decimal(request.POST.get("gst_percent") or "0")
    advance_received = Decimal(request.POST.get("advance_received") or "0")

    # Transport fields
    transport_rate_per_ton = Decimal(request.POST.get("transport_rate_per_ton") or "0")
    transport_paid_by_dealer = Decimal(request.POST.get("transport_paid_by_dealer") or "0")
    transport_paid_by_customer = Decimal(request.POST.get("transport_paid_by_customer") or "0")

    # Internal breakup fields
    purchase_item_ids = request.POST.getlist("purchase_item[]")
    row_bags_list = request.POST.getlist("row_bags[]")

    # -------------------------
    # Validations
    # -------------------------
    if not sale_date or not customer_name:
        messages.error(request, "Sale Date and Customer Name are required.")
        return render(request, "core/add_sale.html", {
            "products": products,
            "brokers": brokers,
            "purchase_items_json": purchase_items_json,
        })

    if product_id <= 0:
        messages.error(request, "Please select a Product.")
        return render(request, "core/add_sale.html", {
            "products": products,
            "brokers": brokers,
            "purchase_items_json": purchase_items_json,
        })

    if total_bags <= 0 or bag_weight <= 0:
        messages.error(request, "Total Bags and Bag Weight must be greater than 0.")
        return render(request, "core/add_sale.html", {
            "products": products,
            "brokers": brokers,
            "purchase_items_json": purchase_items_json,
        })

    if len(purchase_item_ids) != len(row_bags_list):
        messages.error(request, "Internal breakup rows are invalid.")
        return render(request, "core/add_sale.html", {
            "products": products,
            "brokers": brokers,
            "purchase_items_json": purchase_items_json,
        })

    # count only valid rows (pid + bags>0)
    valid_rows = 0
    breakup_sum = 0
    for pid, bags_str in zip(purchase_item_ids, row_bags_list):
        bags = int(bags_str or 0)
        if pid and bags > 0:
            valid_rows += 1
            breakup_sum += bags

    if valid_rows == 0:
        messages.error(request, "Please add at least 1 breakup row (Purchase Stock + Bags).")
        return render(request, "core/add_sale.html", {
            "products": products,
            "brokers": brokers,
            "purchase_items_json": purchase_items_json,
        })

    if breakup_sum != total_bags:
        messages.error(request, f"Internal breakup bags ({breakup_sum}) must match Total Bags ({total_bags}).")
        return render(request, "core/add_sale.html", {
            "products": products,
            "brokers": brokers,
            "purchase_items_json": purchase_items_json,
        })

    # -------------------------
    # Calculations
    # -------------------------
    total_kg = Decimal(total_bags) * bag_weight

    taxable_amount = total_kg * rate_per_kg
    gst_amount = (taxable_amount * gst_percent) / Decimal("100")
    rice_total = taxable_amount + gst_amount

    rice_due = rice_total - advance_received
    if rice_due < 0:
        rice_due = Decimal("0")

    total_ton = total_kg / Decimal("1000")
    transport_charge = total_ton * transport_rate_per_ton

    transport_due = transport_charge - (transport_paid_by_dealer + transport_paid_by_customer)
    if transport_due < 0:
        transport_due = Decimal("0")

    grand_total = rice_total + transport_charge

    # -------------------------
    # Build breakup rows (for review + save)
    # -------------------------
    selected_ids = [int(pid) for pid, bags_str in zip(purchase_item_ids, row_bags_list) if pid and int(bags_str or 0) > 0]

    pi_map = {
        pi.id: pi
        for pi in PurchaseItem.objects.select_related("purchase", "purchase__mill", "product").filter(id__in=selected_ids)
    }

    breakup_rows = []
    buy_cost_total = Decimal("0")

    for pid, bags_str in zip(purchase_item_ids, row_bags_list):
        if not pid:
            continue
        bags = int(bags_str or 0)
        if bags <= 0:
            continue

        pi = pi_map.get(int(pid))
        if not pi:
            continue

        if not pi.bag_weight:
            messages.error(request, "Selected purchase stock has no bag weight.")
            return render(request, "core/add_sale.html", {
                "products": products,
                "brokers": brokers,
                "purchase_items_json": purchase_items_json,
            })

        bw = Decimal(str(pi.bag_weight))
        row_kg = bw * Decimal(bags)
        buy_rate = Decimal(str(pi.purchase_price or 0))
        row_amount = row_kg * buy_rate

        buy_cost_total += row_amount

        breakup_rows.append({
            "purchase_item_id": pi.id,
            "invoice_no": pi.purchase.invoice_no,
            "mill_name": pi.purchase.mill.mill_name,
            "mill_id": pi.purchase.mill_id,
            "product_name": pi.product.rice_name,
            "product_id": pi.product_id,
            "bag_weight": int(pi.bag_weight),
            "bags": bags,
            "kg": row_kg,
            "buy_rate": buy_rate,
            "amount": row_amount,
        })

    profit_estimate = rice_total - buy_cost_total

    # -------------------------
    # REVIEW (no save)
    # -------------------------
    if step == "review":
        return render(request, "core/sale_review.html", {
            "sale_date": sale_date,
            "customer_name": customer_name,
            "customer_gst": customer_gst,
            "broker_id": broker_id,

            "vehicle_number": vehicle_number,
            "driver_name": driver_name,
            "transporter_name": transporter_name,

            "product_id": product_id,
            "bag_weight": bag_weight,
            "total_bags": total_bags,
            "rate_per_kg": rate_per_kg,
            "gst_percent": gst_percent,
            "advance_received": advance_received,

            "total_kg": total_kg,
            "taxable_amount": taxable_amount,
            "gst_amount": gst_amount,
            "rice_total": rice_total,
            "rice_due": rice_due,

            "transport_rate_per_ton": transport_rate_per_ton,
            "transport_paid_by_dealer": transport_paid_by_dealer,
            "transport_paid_by_customer": transport_paid_by_customer,
            "transport_charge": transport_charge,
            "transport_due": transport_due,

            "grand_total": grand_total,
            "breakup_rows": breakup_rows,
            "buy_cost_total": buy_cost_total,
            "profit_estimate": profit_estimate,

            # used for hidden inputs on confirm
            "raw_post": request.POST,
        })

    # -------------------------
    # SAVE (confirm)
    # -------------------------
    invoice_no = generate_sale_invoice_no()

    with transaction.atomic():
        sale = Sale.objects.create(
            invoice_no=invoice_no,

            customer_name=customer_name,
            customer_gst=customer_gst,
            broker_id=broker_id,

            sale_date=sale_date,
            vehicle_number=vehicle_number,
            driver_name=driver_name,
            transporter_name=transporter_name,

            transport_charge=transport_charge,
            transport_paid_by_dealer=transport_paid_by_dealer,
            transport_paid_by_customer=transport_paid_by_customer,

            total_quantity_kg=total_kg,
            taxable_amount=taxable_amount,
            gst_percent=gst_percent,
            gst_amount=gst_amount,

            total_amount=grand_total,

            advance_received=advance_received,
            balance_amount=rice_due,
        )

        # ✅ Save breakup as BUY COST rows (NO NULL fields)
        for row in breakup_rows:
            SaleItem.objects.create(
                sale=sale,
                product_id=row["product_id"],
                mill_id=row["mill_id"],
                bag_weight=row["bag_weight"],
                bag_count=row["bags"],
                rate_per_kg=row["buy_rate"],
                total_weight=row["kg"],
                amount=row["amount"],
            )

    messages.success(request, f"✅ Sale saved: {invoice_no}")
    return redirect("sale_detail", sale_id=sale.id)





# def _d(v, default="0"):
#     """Safe Decimal conversion."""
#     try:
#         if v is None or str(v).strip() == "":
#             return Decimal(default)
#         return Decimal(str(v).strip())
#     except (InvalidOperation, ValueError, TypeError):
#         return Decimal(default)


# def add_sale(request):
    products = Product.objects.filter(is_active=True).order_by("rice_name")
    brokers = Broker.objects.all().order_by("broker_name")

    purchase_items = (
        PurchaseItem.objects
        .select_related("purchase", "purchase__mill", "product")
        .order_by("-purchase__purchase_date", "-id")
    )

    purchase_items_json = json.dumps([
        {
            "id": pi.id,
            "product_id": pi.product_id,
            "bag_weight": pi.bag_weight,
            "label": f"{pi.purchase.invoice_no} / {pi.purchase.mill.mill_name} / Buy ₹{pi.purchase_price}/KG / {pi.bag_weight}kg bag",
        }
        for pi in purchase_items
    ])

    # -------------------------
    # GET → show form (prefill from session draft)
    # -------------------------
    if request.method == "GET":
        draft = request.session.get("sale_draft") or {}
        draft_lists = request.session.get("sale_draft_lists") or {}

        return render(request, "core/add_sale.html", {
            "products": products,
            "brokers": brokers,
            "purchase_items_json": purchase_items_json,

            # ✅ Prefill
            "draft": draft,
            "draft_purchase_items": draft_lists.get("purchase_item", []),
            "draft_row_bags": draft_lists.get("row_bags", []),
        })

    # -------------------------
    # POST
    # -------------------------
    step = request.POST.get("step", "").strip()

    # ✅ Step 2: Confirm & Save (from SESSION, not from hidden inputs)
    if step == "save":
        draft = request.session.get("sale_draft") or {}
        draft_lists = request.session.get("sale_draft_lists") or {}

        if not draft:
            messages.error(request, "Draft not found. Please fill the sale form again.")
            return redirect("add_sale")

        # -------- Read scalar fields from draft --------
        sale_date = draft.get("sale_date") or str(timezone.now().date())
        customer_name = (draft.get("customer_name") or "").strip()
        customer_gst = (draft.get("customer_gst") or "").strip()
        broker_id = draft.get("broker_id") or None

        vehicle_number = (draft.get("vehicle_number") or "").strip()
        driver_name = (draft.get("driver_name") or "").strip()
        transporter_name = (draft.get("transporter_name") or "").strip()

        product_id = draft.get("product_id") or None
        bag_weight = int(draft.get("bag_weight") or 0)
        total_bags = int(draft.get("total_bags") or 0)

        rate_per_kg = _d(draft.get("rate_per_kg"), "0")
        gst_percent = int(draft.get("gst_percent") or 0)
        advance_received = _d(draft.get("advance_received"), "0")

        transport_rate_per_ton = _d(draft.get("transport_rate_per_ton"), "0")
        transport_paid_by_dealer = _d(draft.get("transport_paid_by_dealer"), "0")
        transport_paid_by_customer = _d(draft.get("transport_paid_by_customer"), "0")

        # -------- Basic validation --------
        if not customer_name or not product_id or total_bags <= 0 or bag_weight <= 0:
            messages.error(request, "Missing required fields in draft. Please edit and try again.")
            return redirect("add_sale")

        # -------- Compute totals --------
        total_kg = Decimal(total_bags) * Decimal(bag_weight)
        taxable_amount = total_kg * rate_per_kg
        gst_amount = (taxable_amount * Decimal(gst_percent)) / Decimal("100")
        rice_total = taxable_amount + gst_amount

        rice_due = rice_total - advance_received

        # Transport charge based on ton: (kg / 1000) * rate_per_ton
        transport_charge = (total_kg / Decimal("1000")) * transport_rate_per_ton
        transport_due = transport_charge - (transport_paid_by_dealer + transport_paid_by_customer)

        grand_total = rice_total + transport_charge

        # -------- Breakup arrays from session --------
        purchase_item_ids = draft_lists.get("purchase_item", [])
        row_bags_list = draft_lists.get("row_bags", [])

        breakup_rows = []
        buy_cost_total = Decimal("0")

        # Validate arrays length
        if len(purchase_item_ids) != len(row_bags_list):
            messages.error(request, "Breakup rows mismatch. Please edit and try again.")
            return redirect("add_sale")

        # Build breakup rows and compute buy cost
        for pid, bags_str in zip(purchase_item_ids, row_bags_list):
            bags = int(bags_str or 0)
            if bags <= 0:
                continue

            pi = PurchaseItem.objects.select_related("purchase", "purchase__mill").get(id=int(pid))
            kg = Decimal(bags) * Decimal(pi.bag_weight)
            amount = kg * pi.purchase_price

            buy_cost_total += amount
            breakup_rows.append({
                "purchase_item": pi,
                "purchase_item_id": pi.id,
                "invoice_no": pi.purchase.invoice_no,
                "mill_name": pi.purchase.mill.mill_name,
                "bag_weight": pi.bag_weight,
                "bags": bags,
                "kg": kg,
                "buy_rate": pi.purchase_price,
                "amount": amount,
            })

        profit_estimate = rice_total - buy_cost_total

        # -------- Save Sale --------
        broker = Broker.objects.filter(id=broker_id).first() if broker_id else None
        product = Product.objects.get(id=int(product_id))

        sale = Sale.objects.create(
            invoice_no=f"SALE-{timezone.now().strftime('%Y%m%d%H%M%S')}",
            customer_name=customer_name,
            customer_gst=customer_gst,
            sale_date=sale_date,
            vehicle_number=vehicle_number,
            driver_name=driver_name,
            transporter_name=transporter_name,

            total_quantity_kg=total_kg,
            taxable_amount=taxable_amount,
            gst_percent=gst_percent,
            gst_amount=gst_amount,

            # ✅ Keep this as RICE total (recommended)
            total_amount=rice_total,

            advance_received=advance_received,
            balance_amount=rice_due,

            broker=broker,

            # Transport separate
            transport_charge=transport_charge,
            paid_by_dealer=transport_paid_by_dealer,
            paid_by_customer=transport_paid_by_customer,
        )

        # Save SaleItems breakup rows
        # (This assumes your SaleItem has these fields, adjust if different)
        for r in breakup_rows:
            SaleItem.objects.create(
                sale=sale,
                product=product,
                mill=r["purchase_item"].purchase.mill,
                bag_weight=r["bag_weight"],
                bag_count=r["bags"],
                rate_per_kg=rate_per_kg,  # selling rate
                total_weight=r["kg"],
                amount=(r["kg"] * rate_per_kg),
            )

        # ✅ Clear draft after successful save
        request.session.pop("sale_draft", None)
        request.session.pop("sale_draft_lists", None)
        request.session.modified = True

        messages.success(request, "Sale saved successfully ✅")
        return redirect("sale_detail", sale.id)

    # ✅ Step 1: Review (store draft in session + render review)
    # This runs when POST comes from the form and step != save

    # Save draft to session
    request.session["sale_draft"] = dict(request.POST.items())
    request.session["sale_draft_lists"] = {
        "purchase_item": request.POST.getlist("purchase_item[]"),
        "row_bags": request.POST.getlist("row_bags[]"),
    }
    request.session.modified = True

    # Build review page context from POST (same as before)
    sale_date = request.POST.get("sale_date") or str(timezone.now().date())
    customer_name = request.POST.get("customer_name", "")
    customer_gst = request.POST.get("customer_gst", "")
    broker_id = request.POST.get("broker_id") or ""

    vehicle_number = request.POST.get("vehicle_number", "")
    driver_name = request.POST.get("driver_name", "")
    transporter_name = request.POST.get("transporter_name", "")

    product_id = request.POST.get("product_id") or ""
    bag_weight = int(request.POST.get("bag_weight") or 0)
    total_bags = int(request.POST.get("total_bags") or 0)

    rate_per_kg = _d(request.POST.get("rate_per_kg"), "0")
    gst_percent = int(request.POST.get("gst_percent") or 0)
    advance_received = _d(request.POST.get("advance_received"), "0")

    transport_rate_per_ton = _d(request.POST.get("transport_rate_per_ton"), "0")
    transport_paid_by_dealer = _d(request.POST.get("transport_paid_by_dealer"), "0")
    transport_paid_by_customer = _d(request.POST.get("transport_paid_by_customer"), "0")

    total_kg = Decimal(total_bags) * Decimal(bag_weight)
    taxable_amount = total_kg * rate_per_kg
    gst_amount = (taxable_amount * Decimal(gst_percent)) / Decimal("100")
    rice_total = taxable_amount + gst_amount
    rice_due = rice_total - advance_received

    transport_charge = (total_kg / Decimal("1000")) * transport_rate_per_ton
    transport_due = transport_charge - (transport_paid_by_dealer + transport_paid_by_customer)

    grand_total = rice_total + transport_charge

    purchase_item_ids = request.POST.getlist("purchase_item[]")
    row_bags_list = request.POST.getlist("row_bags[]")

    breakup_rows = []
    buy_cost_total = Decimal("0")

    if len(purchase_item_ids) == len(row_bags_list):
        for pid, bags_str in zip(purchase_item_ids, row_bags_list):
            bags = int(bags_str or 0)
            if bags <= 0:
                continue
            pi = PurchaseItem.objects.select_related("purchase", "purchase__mill").get(id=int(pid))
            kg = Decimal(bags) * Decimal(pi.bag_weight)
            amount = kg * pi.purchase_price
            buy_cost_total += amount
            breakup_rows.append({
                "purchase_item_id": pi.id,
                "invoice_no": pi.purchase.invoice_no,
                "mill_name": pi.purchase.mill.mill_name,
                "bag_weight": pi.bag_weight,
                "bags": bags,
                "kg": kg,
                "buy_rate": pi.purchase_price,
                "amount": amount,
            })

    profit_estimate = rice_total - buy_cost_total

    return render(request, "core/sale_review.html", {
        "sale_date": sale_date,
        "customer_name": customer_name,
        "customer_gst": customer_gst,
        "broker_id": broker_id,

        "vehicle_number": vehicle_number,
        "driver_name": driver_name,
        "transporter_name": transporter_name,

        "product_id": product_id,
        "bag_weight": bag_weight,
        "total_bags": total_bags,
        "total_kg": total_kg,
        "rate_per_kg": rate_per_kg,

        "gst_percent": gst_percent,
        "taxable_amount": taxable_amount,
        "gst_amount": gst_amount,
        "rice_total": rice_total,
        "advance_received": advance_received,
        "rice_due": rice_due,

        "transport_rate_per_ton": transport_rate_per_ton,
        "transport_charge": transport_charge,
        "transport_paid_by_dealer": transport_paid_by_dealer,
        "transport_paid_by_customer": transport_paid_by_customer,
        "transport_due": transport_due,

        "breakup_rows": breakup_rows,
        "buy_cost_total": buy_cost_total,
        "profit_estimate": profit_estimate,
        "grand_total": grand_total,
    })



def _d(v, default="0"):
    """Safe Decimal conversion."""
    try:
        if v is None or str(v).strip() == "":
            return Decimal(default)
        return Decimal(str(v).strip())
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(default)


def add_sale(request):
    products = Product.objects.filter(is_active=True).order_by("rice_name")
    brokers = Broker.objects.all().order_by("broker_name")

    purchase_items = (
        PurchaseItem.objects
        .select_related("purchase", "purchase__mill", "product")
        .order_by("-purchase__purchase_date", "-id")
    )

    purchase_items_json = json.dumps([
        {
            "id": pi.id,
            "product_id": pi.product_id,
            "bag_weight": pi.bag_weight,
            "label": f"{pi.purchase.invoice_no} / {pi.purchase.mill.mill_name} / Buy ₹{pi.purchase_price}/KG / {pi.bag_weight}kg bag",
        }
        for pi in purchase_items
    ])

    # -------------------------
    # GET → show form (prefill from session draft)
    # -------------------------
    if request.method == "GET":
        draft = request.session.get("sale_draft") or {}
        draft_lists = request.session.get("sale_draft_lists") or {}

        return render(request, "core/add_sale.html", {
            "products": products,
            "brokers": brokers,
            "purchase_items_json": purchase_items_json,

            # ✅ Prefill
            "draft": draft,
            "draft_purchase_items": draft_lists.get("purchase_item", []),
            "draft_row_bags": draft_lists.get("row_bags", []),
        })

    # -------------------------
    # POST
    # -------------------------
    step = (request.POST.get("step") or "").strip()

    # ✅ Step 2: Confirm & Save (from SESSION)
    if step == "save":
        draft = request.session.get("sale_draft") or {}
        draft_lists = request.session.get("sale_draft_lists") or {}

        if not draft:
            messages.error(request, "Draft not found. Please fill the sale form again.")
            return redirect("add_sale")

        # -------- Read scalar fields from draft --------
        sale_date = draft.get("sale_date") or str(timezone.now().date())
        customer_name = (draft.get("customer_name") or "").strip()
        customer_gst = (draft.get("customer_gst") or "").strip()
        broker_id = draft.get("broker_id") or None

        vehicle_number = (draft.get("vehicle_number") or "").strip()
        driver_name = (draft.get("driver_name") or "").strip()
        transporter_name = (draft.get("transporter_name") or "").strip()

        product_id = draft.get("product_id") or None
        bag_weight = int(draft.get("bag_weight") or 0)
        total_bags = int(draft.get("total_bags") or 0)

        sell_rate_per_kg = _d(draft.get("rate_per_kg"), "0")
        gst_percent = _d(draft.get("gst_percent"), "0")
        advance_received = _d(draft.get("advance_received"), "0")

        transport_rate_per_ton = _d(draft.get("transport_rate_per_ton"), "0")
        transport_paid_by_dealer = _d(draft.get("transport_paid_by_dealer"), "0")
        transport_paid_by_customer = _d(draft.get("transport_paid_by_customer"), "0")

        # -------- Validation --------
        if not customer_name:
            messages.error(request, "Customer name is required.")
            return redirect("add_sale")

        if not product_id:
            messages.error(request, "Product is required.")
            return redirect("add_sale")

        if total_bags <= 0 or bag_weight <= 0:
            messages.error(request, "Total bags and bag weight must be greater than 0.")
            return redirect("add_sale")

        # -------- Compute Rice Totals --------
        total_kg = Decimal(total_bags) * Decimal(bag_weight)
        taxable_amount = total_kg * sell_rate_per_kg
        gst_amount = (taxable_amount * gst_percent) / Decimal("100")
        rice_total = taxable_amount + gst_amount
        rice_due = rice_total - advance_received

        # -------- Compute Transport --------
        transport_charge = (total_kg / Decimal("1000")) * transport_rate_per_ton
        transport_due = transport_charge - (transport_paid_by_dealer + transport_paid_by_customer)

        # -------- Breakup arrays from session --------
        purchase_item_ids = draft_lists.get("purchase_item", [])
        row_bags_list = draft_lists.get("row_bags", [])

        # Validate arrays length
        if len(purchase_item_ids) != len(row_bags_list):
            messages.error(request, "Breakup rows mismatch. Please edit and try again.")
            return redirect("add_sale")

        # Validate breakup sum = total_bags (IMPORTANT)
        breakup_sum = 0
        for bags_str in row_bags_list:
            breakup_sum += int(bags_str or 0)

        if breakup_sum != total_bags:
            messages.error(
                request,
                f"Breakup bags total must match Total Bags. (Breakup={breakup_sum}, Total={total_bags})"
            )
            return redirect("add_sale")

        # -------- Build breakup rows (BUY cost) --------
        breakup_rows = []
        buy_cost_total = Decimal("0")

        for pid, bags_str in zip(purchase_item_ids, row_bags_list):
            bags = int(bags_str or 0)
            if bags <= 0:
                continue

            pi = PurchaseItem.objects.select_related("purchase", "purchase__mill").get(id=int(pid))

            # BUY side
            kg = Decimal(bags) * Decimal(pi.bag_weight)
            buy_rate = pi.purchase_price
            buy_amount = kg * buy_rate

            buy_cost_total += buy_amount

            breakup_rows.append({
                "purchase_item": pi,
                "bags": bags,
                "kg": kg,
                "buy_rate": buy_rate,
                "buy_amount": buy_amount,
            })

        # -------- Save Sale --------
        broker = Broker.objects.filter(id=broker_id).first() if broker_id else None
        product = Product.objects.get(id=int(product_id))

        sale = Sale.objects.create(
            invoice_no=f"SALE-{timezone.now().strftime('%Y%m%d%H%M%S')}",
            customer_name=customer_name,
            customer_gst=customer_gst,
            broker=broker,
            sale_date=sale_date,

            vehicle_number=vehicle_number,
            driver_name=driver_name,
            transporter_name=transporter_name,

            # ✅ Transport fields (match your model)
            transport_rate_per_ton=transport_rate_per_ton,
            transport_charge=transport_charge,
            transport_paid_by_dealer=transport_paid_by_dealer,
            transport_paid_by_customer=transport_paid_by_customer,

            # ✅ Rice fields
            total_quantity_kg=total_kg,
            taxable_amount=taxable_amount,
            gst_percent=gst_percent,
            gst_amount=gst_amount,
            total_amount=rice_total,      # rice total only
            advance_received=advance_received,
            balance_amount=rice_due,      # rice due only
        )

        # -------- Save SaleItems = BUY breakup rows --------
        for r in breakup_rows:
            pi = r["purchase_item"]

            SaleItem.objects.create(
                sale=sale,
                product=product,
                mill=pi.purchase.mill,

                bag_weight=pi.bag_weight,
                bag_count=r["bags"],

                # ✅ This model stores BUY values
                rate_per_kg=r["buy_rate"],
                total_weight=r["kg"],
                amount=r["buy_amount"],
            )

        # ✅ Clear draft after successful save
        request.session.pop("sale_draft", None)
        request.session.pop("sale_draft_lists", None)
        request.session.modified = True

        messages.success(request, "Sale saved successfully ✅")
        return redirect("sale_detail", sale.id)

    # ✅ Step 1: Review (store draft in session + render review)
    # Save draft to session
    request.session["sale_draft"] = dict(request.POST.items())
    request.session["sale_draft_lists"] = {
        "purchase_item": request.POST.getlist("purchase_item[]"),
        "row_bags": request.POST.getlist("row_bags[]"),
    }
    request.session.modified = True

    # Build review page context from POST
    sale_date = request.POST.get("sale_date") or str(timezone.now().date())
    customer_name = request.POST.get("customer_name", "")
    customer_gst = request.POST.get("customer_gst", "")
    broker_id = request.POST.get("broker_id") or ""

    vehicle_number = request.POST.get("vehicle_number", "")
    driver_name = request.POST.get("driver_name", "")
    transporter_name = request.POST.get("transporter_name", "")

    product_id = request.POST.get("product_id") or ""
    bag_weight = int(request.POST.get("bag_weight") or 0)
    total_bags = int(request.POST.get("total_bags") or 0)

    sell_rate_per_kg = _d(request.POST.get("rate_per_kg"), "0")
    gst_percent = _d(request.POST.get("gst_percent"), "0")
    advance_received = _d(request.POST.get("advance_received"), "0")

    transport_rate_per_ton = _d(request.POST.get("transport_rate_per_ton"), "0")
    transport_paid_by_dealer = _d(request.POST.get("transport_paid_by_dealer"), "0")
    transport_paid_by_customer = _d(request.POST.get("transport_paid_by_customer"), "0")

    total_kg = Decimal(total_bags) * Decimal(bag_weight)
    taxable_amount = total_kg * sell_rate_per_kg
    gst_amount = (taxable_amount * gst_percent) / Decimal("100")
    rice_total = taxable_amount + gst_amount
    rice_due = rice_total - advance_received

    transport_charge = (total_kg / Decimal("1000")) * transport_rate_per_ton
    transport_due = transport_charge - (transport_paid_by_dealer + transport_paid_by_customer)

    grand_total = rice_total + transport_charge

    purchase_item_ids = request.POST.getlist("purchase_item[]")
    row_bags_list = request.POST.getlist("row_bags[]")

    breakup_rows = []
    buy_cost_total = Decimal("0")

    if len(purchase_item_ids) == len(row_bags_list):
        for pid, bags_str in zip(purchase_item_ids, row_bags_list):
            bags = int(bags_str or 0)
            if bags <= 0:
                continue
            pi = PurchaseItem.objects.select_related("purchase", "purchase__mill").get(id=int(pid))
            kg = Decimal(bags) * Decimal(pi.bag_weight)
            buy_amount = kg * pi.purchase_price
            buy_cost_total += buy_amount

            breakup_rows.append({
                "purchase_item_id": pi.id,
                "invoice_no": pi.purchase.invoice_no,
                "mill_name": pi.purchase.mill.mill_name,
                "bag_weight": pi.bag_weight,
                "bags": bags,
                "kg": kg,
                "buy_rate": pi.purchase_price,
                "amount": buy_amount,
            })

    profit_estimate = rice_total - buy_cost_total

    return render(request, "core/sale_review.html", {
        "sale_date": sale_date,
        "customer_name": customer_name,
        "customer_gst": customer_gst,
        "broker_id": broker_id,

        "vehicle_number": vehicle_number,
        "driver_name": driver_name,
        "transporter_name": transporter_name,

        "product_id": product_id,
        "bag_weight": bag_weight,
        "total_bags": total_bags,
        "total_kg": total_kg,
        "rate_per_kg": sell_rate_per_kg,

        "gst_percent": gst_percent,
        "taxable_amount": taxable_amount,
        "gst_amount": gst_amount,
        "rice_total": rice_total,
        "advance_received": advance_received,
        "rice_due": rice_due,

        "transport_rate_per_ton": transport_rate_per_ton,
        "transport_charge": transport_charge,
        "transport_paid_by_dealer": transport_paid_by_dealer,
        "transport_paid_by_customer": transport_paid_by_customer,
        "transport_due": transport_due,

        "breakup_rows": breakup_rows,
        "buy_cost_total": buy_cost_total,
        "profit_estimate": profit_estimate,
        "grand_total": grand_total,
    })

def sale_review(request):
    if request.method != "POST":
        return redirect("add_sale")

    # Get all form data
    data = request.POST.copy()

    bag_weight = Decimal(data.get("bag_weight") or "0")
    total_bags = Decimal(data.get("total_bags") or "0")
    rate_per_kg = Decimal(data.get("rate_per_kg") or "0")
    gst_percent = Decimal(data.get("gst_percent") or "0")
    advance = Decimal(data.get("advance_received") or "0")

    total_kg = bag_weight * total_bags
    taxable = total_kg * rate_per_kg
    gst_amt = (taxable * gst_percent) / Decimal("100")
    rice_total = taxable + gst_amt
    rice_due = rice_total - advance

    total_ton = total_kg / Decimal("1000")
    transport_rate = Decimal(data.get("transport_rate_per_ton") or "0")
    transport_charge = total_ton * transport_rate

    transport_due = transport_charge - (
        Decimal(data.get("transport_paid_by_dealer") or "0") +
        Decimal(data.get("transport_paid_by_customer") or "0")
    )

    context = {
        "data": data,
        "total_kg": round(total_kg, 2),
        "taxable": round(taxable, 2),
        "gst_amt": round(gst_amt, 2),
        "rice_total": round(rice_total, 2),
        "rice_due": round(rice_due, 2),
        "transport_charge": round(transport_charge, 2),
        "transport_due": round(transport_due, 2),
    }

    return render(request, "core/sale_review.html", context)

@transaction.atomic
def sale_confirm_save(request):
    if request.method != "POST":
        return redirect("sale_list")

    invoice_no = generate_sale_invoice_no()

    sale = Sale.objects.create(
        invoice_no=invoice_no,
        customer_name=request.POST.get("customer_name"),
        customer_gst=request.POST.get("customer_gst"),
        broker_id=request.POST.get("broker_id") or None,
        sale_date=request.POST.get("sale_date"),

        vehicle_number=request.POST.get("vehicle_number"),
        driver_name=request.POST.get("driver_name"),
        transporter_name=request.POST.get("transporter_name"),

        total_quantity_kg=request.POST.get("total_kg"),
        taxable_amount=request.POST.get("taxable"),
        gst_percent=request.POST.get("gst_percent"),
        gst_amount=request.POST.get("gst_amt"),

        transport_charge=request.POST.get("transport_charge"),
        transport_paid_by_dealer=request.POST.get("transport_paid_by_dealer") or 0,
        transport_paid_by_customer=request.POST.get("transport_paid_by_customer") or 0,

        advance_received=request.POST.get("advance_received") or 0,
        total_amount=request.POST.get("rice_total"),
        balance_amount=request.POST.get("rice_due"),
    )

    return redirect("sale_print", sale_id=sale.id)

def sale_print(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id)

    rice_total = float(sale.taxable_amount) + float(sale.gst_amount)

    return render(request, "core/sale_print.html", {
        "sale": sale,
        "rice_total": rice_total,
    })



def broker_list(request):
    q = request.GET.get("q", "").strip()

    brokers = Broker.objects.all().order_by("-created_at")
    if q:
        brokers = brokers.filter(broker_name__icontains=q)

    return render(request, "core/broker_list.html", {
        "brokers": brokers
    })


def add_broker(request):
    if request.method == "POST":
        Broker.objects.create(
            broker_name=request.POST.get("broker_name"),
            mobile=request.POST.get("mobile", ""),
            gst_number=request.POST.get("gst_number", ""),
            opening_balance=request.POST.get("opening_balance") or 0,
            address=request.POST.get("address", "")
        )

        messages.success(request, "✅ Broker saved successfully!")
        return redirect("broker_list")

    return render(request, "core/add_broker.html")


def broker_report_detail(request, broker_id):
    broker = get_object_or_404(Broker, id=broker_id)

    sales = Sale.objects.filter(broker=broker).order_by("-sale_date", "-id")

    total_sales = sales.aggregate(s=Sum("total_amount"))["s"] or 0
    total_advance = sales.aggregate(s=Sum("advance_received"))["s"] or 0
    total_due = float(total_sales) - float(total_advance)

    return render(request, "core/broker_report_detail.html", {
        "broker": broker,
        "sales": sales,
        "total_sales": total_sales,
        "total_advance": total_advance,
        "total_due": total_due
    })


from io import BytesIO
from decimal import Decimal
import qrcode

from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth

from .models import Sale, SaleItem



def _amount_words(n: Decimal):
    number = int(n)
    words = num2words(number, lang='en_IN')
    return words.title() + " Only"


def sale_invoice_pdf(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id)
    items = SaleItem.objects.filter(sale=sale).select_related("product", "mill")

    # ===== Company =====
    company_name = getattr(settings, "COMPANY_NAME", "Company Name")
    company_address = getattr(settings, "COMPANY_ADDRESS", "")
    company_phone = getattr(settings, "COMPANY_PHONE", "")
    company_email = getattr(settings, "COMPANY_EMAIL", "")
    company_gstin = getattr(settings, "COMPANY_GSTIN", "")
    company_pan = getattr(settings, "COMPANY_PAN", "")

    # ===== Bank =====
    bank_ac_name = getattr(settings, "BANK_ACCOUNT_NAME", company_name)
    bank_ac_no = getattr(settings, "BANK_ACCOUNT_NO", "")
    bank_name = getattr(settings, "BANK_NAME", "")
    bank_ifsc = getattr(settings, "BANK_IFSC", "")
    bank_branch = getattr(settings, "BANK_BRANCH", "")
    upi_id = getattr(settings, "UPI_ID", "")

    # ===== Amounts (invoice total = RICE ONLY) =====
    taxable = Decimal(str(sale.taxable_amount or 0))
    gst_amt = Decimal(str(sale.gst_amount or 0))
    gst_percent = Decimal(str(sale.gst_percent or 0))
    rice_total = Decimal(str(sale.total_amount or 0))  # ✅ rice only

    rice_advance = Decimal(str(sale.advance_received or 0))
    rice_due = Decimal(str(sale.balance_amount or 0))

    # ===== Transport (INFO ONLY; not included in invoice total) =====
    total_kg = Decimal(str(sale.total_quantity_kg or 0))
    total_ton = (total_kg / Decimal("1000")) if total_kg else Decimal("0")

    transport_rate = Decimal(str(sale.transport_rate_per_ton or 0))
    transport_amt = Decimal(str(sale.transport_charge or 0))
    paid_dealer = Decimal(str(sale.transport_paid_by_dealer or 0))
    paid_customer = Decimal(str(sale.transport_paid_by_customer or 0))
    transport_due = transport_amt - (paid_dealer + paid_customer)
    if transport_due < 0:
        transport_due = Decimal("0")

    # ===== Product line (one product per sale) =====
    first = items.first()
    product_name = first.product.rice_name if first else "Rice Sale"
    hsn = first.product.hsn_code if (first and first.product.hsn_code) else "-"
    total_bags = sum(int(x.bag_count or 0) for x in items) if first else 0

    sell_rate = (taxable / total_kg) if total_kg else Decimal("0")

    # ===== QR (OPTIONAL) =====
    qr_reader = None
    try:
        qr_text = (
            f"TAX INVOICE\n"
            f"Invoice: {sale.invoice_no}\n"
            f"Date: {sale.sale_date}\n"
            f"Customer: {sale.customer_name}\n"
            f"Rice Total: {rice_total}\n"
            f"Rice Due: {rice_due}\n"
        )
        qr = qrcode.QRCode(box_size=5, border=2)
        qr.add_data(qr_text)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_buf = BytesIO()
        qr_img.save(qr_buf, format="PNG")
        qr_buf.seek(0)
        qr_reader = ImageReader(qr_buf)
    except Exception:
        qr_reader = None

    # ===== PDF =====
    out = BytesIO()
    c = canvas.Canvas(out, pagesize=A4)
    W, H = A4

    L = 12 * mm
    R = W - 12 * mm
    TOP = H - 12 * mm
    BOT = 12 * mm
    BW = R - L

    PAD = 3 * mm
    LINE = 1

    def rect(x, y, w, h, lw=LINE):
        c.setLineWidth(lw)
        c.rect(x, y, w, h)

    def vline(x, y1, y2, lw=LINE):
        c.setLineWidth(lw)
        c.line(x, y1, x, y2)

    def hline(x1, x2, y, lw=LINE):
        c.setLineWidth(lw)
        c.line(x1, y, x2, y)

    def txt(x, y, s, size=9, bold=False):
        c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        c.drawString(x, y, str(s))

    def rtxt(x, y, s, size=9, bold=False):
        c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        c.drawRightString(x, y, str(s))

    def ctxt(x, y, s, size=10, bold=False):
        c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        c.drawCentredString(x, y, str(s))

    def fit_text(s, max_w, size=9, bold=False):
        """Truncate text to fit max width (prevents overlap in columns)."""
        font = "Helvetica-Bold" if bold else "Helvetica"
        s = str(s)
        if stringWidth(s, font, size) <= max_w:
            return s
        ell = "..."
        while s and stringWidth(s + ell, font, size) > max_w:
            s = s[:-1]
        return (s + ell) if s else ell

    def wrap_lines(s, max_w, size=8, bold=False):
        """Word-wrap to fit inside a column width."""
        font = "Helvetica-Bold" if bold else "Helvetica"
        words = str(s).split()
        lines, line = [], ""
        for w in words:
            test = (line + " " + w).strip()
            if stringWidth(test, font, size) <= max_w:
                line = test
            else:
                if line:
                    lines.append(line)
                line = w
        if line:
            lines.append(line)
        return lines

    y = TOP

    # ===== Top strip (padding) =====
    strip_h = 12 * mm
    rect(L, y - strip_h, BW, strip_h)
    txt(L + PAD, y - 8.5, "Page No. 1 of 1", size=9)
    ctxt(L + BW / 2, y - 8.5, "Bill of Supply", size=10, bold=True)
    rtxt(R - PAD, y - 8.5, "Original Copy", size=9)
    y -= strip_h

    # ===== Company box =====
    comp_h = 34 * mm
    rect(L, y - comp_h, BW, comp_h)
    ctxt(L + BW/2, y - 12, company_name, size=12, bold=True)
    ctxt(L + BW/2, y - 24, company_address, size=9)
    ctxt(L + BW/2, y - 34, f"Mobile: +91 {company_phone} | Email: {company_email}", size=9)

    gstpan = " | ".join([x for x in [
        f"GSTIN - {company_gstin}" if company_gstin else "",
        f"PAN - {company_pan}" if company_pan else ""
    ] if x])
    if gstpan:
        ctxt(L + BW/2, y - 44, gstpan, size=9)

    y -= comp_h

    # ===== Invoice details + Transport =====
    info_h = 42 * mm
    rect(L, y - info_h, BW, info_h)
    mid = L + BW/2
    vline(mid, y - info_h, y)

    txt(L + PAD, y - 10, "Invoice Details", bold=True)
    left_lines = [
        ("Invoice Number", sale.invoice_no),
        ("Invoice Date", str(sale.sale_date)),
        ("Due Date", "-"),
        ("Place of Supply", "-"),
        ("Broker Name", sale.broker.broker_name if sale.broker else "NA"),
    ]
    yy = y - 20
    for k, v in left_lines:
        txt(L + PAD, yy, k, bold=True, size=9)
        txt(L + 55*mm, yy, f": {v}", size=9)
        yy -= 5.5 * mm

    txt(mid + PAD, y - 10, "Transport Details (Info Only)", bold=True)
    right_lines = [
        ("Transporter", sale.transporter_name),
        ("Vehicle No.", f"{sale.vehicle_number}  | {sale.driver_name}"),
        ("Rate/Ton", f"Rs. {transport_rate:.2f}"),
        ("Total Ton", f"{total_ton:.3f}"),
        ("Transport Amount", f"Rs. {transport_amt:.2f}"),
        ("Advance (Dealer)", f"Rs. {paid_dealer:.2f}"),
    ]
    if paid_customer > 0:
        right_lines.append(("Paid (Customer)", f"Rs. {paid_customer:.2f}"))
    right_lines.append(("Transport Due", f"Rs. {transport_due:.2f}"))

    yy = y - 20
    for k, v in right_lines:
        txt(mid + PAD, yy, k, bold=True, size=8)
        txt(mid + 55*mm, yy, f": {v}", size=8)
        yy -= 5.2 * mm

    y -= info_h

    # ===== Billing + Shipping =====
    bs_h = 34 * mm
    rect(L, y - bs_h, BW, bs_h)
    vline(mid, y - bs_h, y)

    txt(L + PAD, y - 10, "Billing Details", bold=True)
    b_lines = [("Name", sale.customer_name), ("GSTIN", sale.customer_gst or "-"), ("Address", "-")]
    yy = y - 22
    for k, v in b_lines:
        txt(L + PAD, yy, k, bold=True)
        txt(L + 45*mm, yy, f": {v}")
        yy -= 6 * mm

    txt(mid + PAD, y - 10, "Shipping Details", bold=True)
    s_lines = [("Name", sale.customer_name), ("GSTIN", sale.customer_gst or "-"), ("Address", "-")]
    yy = y - 22
    for k, v in s_lines:
        txt(mid + PAD, yy, k, bold=True)
        txt(mid + 45*mm, yy, f": {v}")
        yy -= 6 * mm

    y -= bs_h

    # ===== Items table (FIX: columns inside BW so headers never overlap) =====
    table_h = 82 * mm
    rect(L, y - table_h, BW, table_h)

    # ✅ A4 usable width BW = (R-L). Build columns using widths that sum to BW (=186mm)
    # columns: Sr | Description | HSN | Bags | KG | Rate | Taxable | GST% | GST | Amount
    w_sr      = 10 * mm
    w_desc    = 64 * mm
    w_hsn     = 14 * mm
    w_bags    = 12 * mm
    w_kg      = 16 * mm
    w_rate    = 14 * mm
    w_taxable = 18 * mm
    w_gstp    = 10 * mm
    w_gst     = 12 * mm
    w_amount  = 16 * mm

    # Sanity: (w_sr+w_desc+...+w_amount) == BW
    col = [L]
    for w in [w_sr, w_desc, w_hsn, w_bags, w_kg, w_rate, w_taxable, w_gstp, w_gst, w_amount]:
        col.append(col[-1] + w)
    # col[-1] should be == R (or extremely close due to float)

    # Draw vertical lines (inside the box)
    for x in col[1:-1]:
        vline(x, y - table_h, y)

    # Header separator
    header_h = 12 * mm
    hline(L, R, y - header_h)

    # ---- Header row (small font for tight columns) ----
    txt(L + 2,      y - 9*mm, "Sr", bold=True, size=8)
    txt(col[1] + 2, y - 9*mm, "Item Description", bold=True, size=8)
    txt(col[2] + 2, y - 9*mm, "HSN", bold=True, size=8)

    rtxt(col[4] - 2, y - 9*mm, "Bags",   bold=True, size=8)
    rtxt(col[5] - 2, y - 9*mm, "KG",     bold=True, size=8)
    rtxt(col[6] - 2, y - 9*mm, "Rate",   bold=True, size=8)
    rtxt(col[7] - 2, y - 9*mm, "Taxable",bold=True, size=8)
    rtxt(col[8] - 2, y - 9*mm, "GST%",   bold=True, size=8)
    rtxt(col[9] - 2, y - 9*mm, "GST",    bold=True, size=8)
    rtxt(R - 2,      y - 9*mm, "Amount", bold=True, size=8)

    # ---- Data row ----
    row_y = y - 22 * mm
    txt(L + 2, row_y, "1", size=9)

    # Description must never enter HSN column
    desc_max_w = (col[2] - col[1]) - 6
    safe_desc = fit_text(product_name, desc_max_w, size=9)
    txt(col[1] + 2, row_y, safe_desc, size=9)

    txt(col[2] + 2, row_y, hsn, size=9)

    rtxt(col[4] - 2, row_y, str(total_bags),     size=9)
    rtxt(col[5] - 2, row_y, f"{total_kg:.2f}",   size=9)
    rtxt(col[6] - 2, row_y, f"{sell_rate:.2f}",  size=9)
    rtxt(col[7] - 2, row_y, f"{taxable:.2f}",    size=9)
    rtxt(col[8] - 2, row_y, f"{gst_percent:.2f}",size=9)
    rtxt(col[9] - 2, row_y, f"{gst_amt:.2f}",    size=9)
    rtxt(R - 2,      row_y, f"{rice_total:.2f}", size=9)

    y -= table_h

    # ===== Rice totals =====
    tot_h = 38 * mm   # 🔥 reduce from 44mm to 38mm
    rect(L, y - tot_h, BW, tot_h)

    txt(L + PAD, y - 10, "Rice Payment Summary (Invoice Total = Rice Only)", bold=True)

    txt(L + PAD, y - 20, "Taxable Amount", bold=True)
    txt(L + 52*mm, y - 20, f": Rs. {taxable:.2f}")

    txt(L + PAD, y - 28, "GST Amount", bold=True)
    txt(L + 52*mm, y - 28, f": Rs. {gst_amt:.2f}")

    txt(L + PAD, y - 36, "Advance Received (Rice)", bold=True)
    txt(L + 52*mm, y - 36, f": Rs. {rice_advance:.2f}")

    rtxt(R - PAD, y - 20, f"Rice Total: Rs. {rice_total:.2f}", bold=True)
    rtxt(R - PAD, y - 30, f"Rice Due: Rs. {rice_due:.2f}", bold=True)

    txt(L + PAD, y - 46, f"Amount in Words: {_amount_words(rice_total)}", size=9, bold=True)

    y -= tot_h

    # ===== Bottom section: Terms | Bank | QR/Stamp =====
    bottom_h = y - BOT
    rect(L, BOT, BW, bottom_h)

    c1 = L + BW/3
    c2 = L + 2*BW/3
    vline(c1, BOT, y)
    vline(c2, BOT, y)

    # Terms
    txt(L + PAD, y - 12, "Terms and Conditions", bold=True, size=10)

    terms = [
        "E & O.E.",
        "1. Goods once sold will not be taken back.",
        "2. Interest @ 18% p.a. will be charged if payment is delayed.",
        "3. In case of non-payment, legal action may be initiated.",
        "4. Subject to local jurisdiction only."
    ]

    terms_left = L + PAD
    terms_right = c1 - PAD
    terms_w = terms_right - terms_left

    ty = y - 24   # 🔥 little more space from title

    for t in terms:
        lines = wrap_lines(t, terms_w, size=8)
        for line in lines:
            if ty < BOT + 20:   # 🔥 increase bottom margin safety
                break
            txt(terms_left, ty, line, size=8)
            ty -= 9  # 🔥 slightly tighter spacing so all lines fit



    # Bank details
    midL = c1
    txt(midL + PAD, y - 12, "Bank / Payment Details", bold=True, size=10)
    by = y - 26
    txt(midL + PAD, by, f"A/C Name: {bank_ac_name}", size=8); by -= 10
    txt(midL + PAD, by, f"A/C No: {bank_ac_no}", size=8); by -= 10
    txt(midL + PAD, by, f"Bank: {bank_name}", size=8); by -= 10
    txt(midL + PAD, by, f"IFSC: {bank_ifsc}", size=8); by -= 10
    txt(midL + PAD, by, f"Branch: {bank_branch}", size=8); by -= 10
    if upi_id:
        txt(midL + PAD, by, f"UPI: {upi_id}", size=8)

    # Right: QR or Stamp
    rightL = c2
    txt(rightL + PAD, y - 12, "Stamp / QR", bold=True, size=10)

    stamp_box = 62 * mm
    box_x = rightL + (BW/3 - stamp_box)/2
    box_y = y - 80 * mm

    rect(box_x, box_y, stamp_box, stamp_box)

    if qr_reader is not None and box_y > (BOT + 5*mm):
        c.drawImage(qr_reader, box_x + 2, box_y + 2, width=stamp_box-4, height=stamp_box-4, mask="auto")
        ctxt(box_x + stamp_box/2, box_y + stamp_box + 3, "QR", size=9, bold=True)

    rtxt(R - PAD, BOT + 25, f"For {company_name}", size=9)
    rtxt(R - PAD, BOT + 12, "Authorized Signatory", size=9)

    c.showPage()
    c.save()

    pdf = out.getvalue()
    out.close()

    filename = f"Invoice_{sale.invoice_no}.pdf"
    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.write(pdf)
    return resp